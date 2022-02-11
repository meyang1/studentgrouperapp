import openpyxl
import json
from app import studentClass


class Class:

    def __init__(self):
        """
        Creates a class that holds, loads, and manipulates student objects
        """
        
        self.students = []

    def getNames(self):
        """
        Returns all the names of the students in the class. Only works after
        .loadFromFile() has been called.
        """

        res=[]
        count=1
        num=1
        rows=self.sheet.max_row 
    
        while num < rows:
            if((rows-num)==1):
                res.append([count, self.sheet.cell(row=num+1, column=1).value, " "])
                num += 1
            else:
                res.append([count, self.sheet.cell(row=num+1, column=1).value, self.sheet.cell(row=num+2, column=1).value])
                num += 2
            count+=1
        
        return res

    def getStudents(self):
        """
        Returns all the students in the class. Only works after .loadFromFile()
        has been called.
        """

        return self.students

    def loadFromString(self, str):
        """
        Converts the string to a dictionary
        """

        try:
            dictionary=json.loads(str)
        except:
            return "Bad string: Can't be converted"

        for key in dictionary:
            self.students.append(studentClass.Student(key, dictionary[key]))

        return self

    def convertToString(self):
        """
        Creates a dictionary of students and converts it to a string
        """
        
        d={}
        for student in self.students:
            d[student.name]=student.cats
        return str(d)

    def loadFromFile(self, path):
        """
        Load the class's students from an Excel file. Needed for most other methods
        to function
        path=path to Excel file
        """
        #print(path)

        try:
            self.sheet = openpyxl.load_workbook(path).active
        except:
            print("Bad Data: File does not exist or is not in current working directory")
            return "Bad Data: File does not exist or is not in current working directory"
            

        rows=self.sheet.max_row
        names=[]

        for i in range(1, rows):
            name=self.sheet.cell(row=i+1, column=1).value
            if name==None:
                print("Bad Data: There is a blank cell in the names column")
                return "Bad Data: There is a blank cell in the names column"
                
            names.append(self.sheet.cell(row=i+1, column=1).value)
        
        for name in names:
            self.students.append(studentClass.Student(name, {}))

        cols=self.sheet.max_column

        for i in range(len(self.students)):
            for j in range(2, cols+1):
                if self.sheet.cell(row=1, column=j).value==None:
                    print("Bad Data: There is a category without a name")
                    return "Bad Data: There is a category without a name"
                    
                if self.sheet.cell(row=i+2, column=j).value==None:
                    print("Bad Data: There is a ranking cell that does not have a proper value")
                    return "Bad Data: There is a ranking cell that does not have a proper value"

                self.students[i].addRank(self.sheet.cell(row=1, column=j).value, self.sheet.cell(row=i+2, column=j).value)

        #return self.students
        return "Good data"

    def group(self, size, sortType, balanced, category=None):
        """
        Puts the students of the class into groups
        sortType values:
        1=Group by average rank
        2=Group by an individual category
        3=Group by getting one person who's good at everything in each group

        balanced values:
        true=The skill level of all of the groups are similar
        false=The students with similar ranks go together

        category: The category the user wants to sort the students
        by. Should only be specified for sortType 2.
        """

        if sortType==1:
            
            averages={}
            scores=[]
            for student in self.students:
                count=0
                for cat in student.cats:
                    count+=int(student.cats[cat])
                scores.append(count)
                averages[student]=count

            scores.sort()
            res=[]

            for score in scores:
                for student in averages:
                    if averages[student]==score:
                        res.append(str(student))
                        break

            if not(balanced):
                ans=[[]]

                count=0
                while len(res)>0:
                    if count<size:
                        ans[-1].append(res.pop())
                        count+=1

                    else:
                        ans.append([])
                        count=0

            else:
                ans=[[]]

                popFromEnd=True

                count=0
                while len(res)>0:
                    if count<size:
                        if popFromEnd:
                            ans[-1].append(res.pop())
                        else:
                            ans[-1].append(res.pop(0))
                        
                        popFromEnd=not(popFromEnd)
                        count+=1

                    else:
                        ans.append([])
                        count=0

            return ans

        elif sortType==2:
            if(category==None):
                print("Sort type 2 requires a category")
                return ["Sort type 2 requires a category"]
            
            if not(category in self.students[0].cats):
                print("The category you entered isn't in the spreadsheet")
                return ["The category you entered isn't in the spreadsheet"]

            scores=[]
            for student in self.students:
                scores.append(student.cats[category])

            scores.sort()
            res=[]

            for score in scores:
                for student in self.students:
                    if student.cats[category]==score:
                        res.append(str(student))
                        break

            if not(balanced):
                ans=[[]]

                count=0
                while len(res)>0:
                    if count<size:
                        ans[-1].append(res.pop())
                        count+=1

                    else:
                        ans.append([])
                        count=0

            else:
                ans=[[]]

                popFromEnd=True

                count=0
                while len(res)>0:
                    if count<size:
                        if popFromEnd:
                            ans[-1].append(res.pop())
                        else:
                            ans[-1].append(res.pop(0))
                        
                        popFromEnd=not(popFromEnd)
                        count+=1

                    else:
                        ans.append([])
                        count=0

            return ans

        elif sortType==3:
            sortedCats=[]

            for cat in self.students[0].cats:
                category=cat
                scores=[]
                for student in self.students:
                    scores.append(student.cats[category])

                scores.sort()
                res=[]

                for score in scores:
                    for student in self.students:
                        if student.cats[category]==score:
                            res.append(str(student))
                            break

                sortedCats.append(res)

            if not(balanced):
                ans=[[]]
                
                index=0
                count=0
                while sortedCats.count([])<len(sortedCats):
                    
                    if len(sortedCats[index])==0:
                        index=(index+1)%len(sortedCats)
                        continue

                    val=sortedCats[index].pop()
                    ans[-1].append(val)
                    index=(index+1)%len(sortedCats)

                    if count<size-1:
                        count+=1
                    
                    else:
                        count=0
                        ans.append([])
                    
                    for arr in sortedCats:                            
                        if val in arr:
                            arr.remove(val)
                return ans

            else:
                print("Sort type 3 can only be unbalanced")
                return ["Sort type 3 can only be unbalanced"]

        else:
            print("Not a valid sort type")
            return["Not a valid sort type"]
    
    def __str__(self):
        """
        Returns a string representation of the class
        """

        res=""
    
        for student in self.students:
            res+=str(student)+"\n" + "\n"

        return res